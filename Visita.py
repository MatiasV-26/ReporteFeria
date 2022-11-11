import json
from openpyxl import workbook, load_workbook

d = open("jsonFeria1111.json", encoding="utf8")
data = json.load(d)
wb = load_workbook("Reporte.xlsx")
ws = wb.active

cont = 0
q = data["Visita"]
lista = []
#Cargar todas las empresas en el Excel
for  i1 in q.values():
    for i2 in i1.values():
        for i3 in i2.values():
            for i4 in i3.values():
                for i5 in i4.values():  
                        e = i5['empresaV']
                        fecha = i5["fecha"]
                        if e not in lista and e != "prueba":
                            lista.append(e)
                            cont += 1
                            ws[f"A{cont + 1}"] = i5['empresaV']
#print(lista)
IngenieriaSistemasCount = 0
adminCount = 0
ComunicaCount = 0
economiaCount = 0
negoCount = 0
derechoCount = 0
industrialCount = 0
civilCount = 0
contaCount = 0
arquiCount = 0
psicoCount = 0
markeCount = 0

listaEmpresas = []
i = data["Visita"]
listaCodigos = []
for  ingreso1 in i.values():
    for ingreso2 in ingreso1.values():
        for ingreso3 in ingreso2.values():
            for ingreso4 in ingreso3.values():
                for ingreso5 in ingreso4.values():  
                        e = ingreso5['empresaV']
                        fecha = ingreso5["fecha"]
                        if e not in listaEmpresas and fecha[0:5] == "10/11":
                            listaEmpresas.append(e)
                            print(e)
#print(listaEmpresas)
c = 0
for E in listaEmpresas:
    
    i = data["Visita"][E]["10"]["11"]
    listaCodigos = []
    total = 0
    IngenieriaSistemasCount = 0
    adminCount = 0
    ComunicaCount = 0
    economiaCount = 0
    negoCount = 0
    derechoCount = 0
    industrialCount = 0
    civilCount = 0
    contaCount = 0
    arquiCount = 0
    psicoCount = 0
    markeCount = 0
    for  ingreso1 in i.values():
        for ingreso2 in ingreso1.values():
            id = ingreso2['idV']
            if id not in listaCodigos:
                listaCodigos.append(id)
                if ingreso2["carreraV"] == "INGENIERÍA DE SISTEMAS":
                    IngenieriaSistemasCount += 1
                elif ingreso2["carreraV"] == "ADMINISTRACIÓN":
                    adminCount += 1
                elif ingreso2["carreraV"] == "ECONOMÍA":
                    economiaCount += 1
                elif ingreso2["carreraV"] == "NEGOCIOS INTERNACIONALES":
                    negoCount += 1
                elif ingreso2["carreraV"] == "DERECHO":
                    derechoCount += 1
                elif ingreso2["carreraV"] == "INGENIERÍA INDUSTRIAL":
                    industrialCount += 1
                elif ingreso2["carreraV"] == "INGENIERÍA CIVIL":
                    civilCount += 1
                elif ingreso2["carreraV"] == "CONTABILIDAD":
                    contaCount += 1
                elif ingreso2["carreraV"] == "COMUNICACIÓN":
                    ComunicaCount += 1
                elif ingreso2["carreraV"] == "ARQUITECTURA":
                    arquiCount += 1
                elif ingreso2["carreraV"] == "PSICOLOGÍA":
                    psicoCount += 1
                elif ingreso2["carreraV"] == "MARKETING":
                    markeCount += 1

    

            #print(ingreso2)
    total = IngenieriaSistemasCount + adminCount + ComunicaCount + economiaCount + negoCount + derechoCount + industrialCount + civilCount + contaCount + arquiCount + psicoCount + markeCount
    c += 1
    #print(c)
    for i in lista:
        a = 0
        for j in range(len(lista)+1):
            emp = ws[f"A{j+1}"]
            a += 1
            if i == ingreso2['empresaV'] and emp.value == ingreso2['empresaV']:
                #print(lista[c], ingreso2['empresaV'])
                ws[f"J{a}"] = total
                a = 0
            
    #print(listaCodigos)
    #print("Visitas a empresa: " + E)
    #print("Visitas totales: " + str(total))
   
#print(cont)  
wb.save("Reporte.xlsx")